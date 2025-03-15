import pandas as pd
import streamlit as st
import os
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.styles.borders import Border, Side
import io

# Функція для обробки файлу
def process_html_file(uploaded_file):

    # Завантаження HTML-файлу
    soup = BeautifulSoup(uploaded_file, "html.parser")

    # Знаходимо таблицю в HTML
    table = soup.find("table")

    # Отримуємо всі рядки таблиці
    rows = table.find_all("tr")

    data = []
    keys_0 = ["No."]
    keys_A = ["Workset A", "Category A", "Family A", "Type A", "ID A"]
    keys_B = ["Model B", "Category B", "Family B", "Type B", "ID B"]

    # Обробка рядків таблиці
    for row in rows[1:]:  # Пропускаємо заголовки

        # Рядок складається з 3 комірок. Розбиваємо по ним.
        raw_cols = row.find_all("td")
        cols = []
        for raw_col in raw_cols:
            cols.append(raw_col.text.strip())
        
        # Перша комірка це лише номер
        separate_cols = {}
        separate_cols[keys_0[0]] = cols[0]

        # Друга комірка це моя модель
        values_A = []
        for item in cols[1].split(" : "):
            values_A.append(item.strip())
        dict_A = dict(zip(keys_A, values_A))
        separate_cols.update(dict_A)

        # Третя комірка це модель колізії
        values_B = []
        for item in cols[2].split(" : "):
            values_B.append(item.strip())

        # виправляємо невірно розбить рядки
        match len(values_B):
            case 7:
                values_B[1] = " : ".join([values_B[1], values_B[2]])
                values_B[3] = " : ".join([values_B[3], values_B[4]])
                del values_B[2]
                del values_B[3]
        dict_B = dict(zip(keys_B, values_B))
        separate_cols.update(dict_B)

        # додаємо стовпчик Знак і видаляємо лишні символи з номеру
        copy_separate_cols = separate_cols.copy()
        for key, value in copy_separate_cols.items():
            if "- Znak" in value:
                type, sign = value.split("- Znak")
                if "A" in key:
                    separate_cols["Type A"] = type.strip()
                    separate_cols["Sign A"] = sign.strip()
                else:
                    separate_cols["Type B"] = type.strip()
                    separate_cols["Sign B"] = sign.strip()
            if "ID " in key:
                separate_cols[key] = value.replace('id ', '').strip()
        data.append(separate_cols)

    df=pd.DataFrame(data, columns=["No.","||","Workset A", "Category A", "Family A", "Type A","Sign A", "ID A","||","Model B", "Category B", "Family B", "Type B","Sign B", "ID B"])

    # Збереження у Excel файл з розширенням .xlsx
    excel_path = "collisions_report.xlsx"
    df.to_excel(excel_path, index=False)

    # Завантажуємо та обробляємо Excel файл
    wb = load_workbook(excel_path)
    ws = wb.active

    # Збереження першого рядка як зафіксованого
    ws.freeze_panes = 'B2'  # заморожує все до другого рядка (включаючи перший)

    # Додаємо фільтри
    ws.auto_filter.ref = ws.dimensions

    # Колір заголовків
    header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    header_font = Font(color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Інші налаштування (наприклад, колір клітинок, бордюри і т.д.)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for idx, cell in enumerate(row):
            if idx % 2 == 0:
                cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    # Tworzenie ramki (Border) dla wszystkich komórek
    border = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )

    # Dodanie ramki do każdej komórki
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = border

    # Автодозор ширини стовпців
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Беремо літеру стовпця
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)  # Додаємо невеликий відступ
        ws.column_dimensions[column].width = adjusted_width   

    # Збереження після внесення змін
    wb.save(excel_path)

    print(f"Excel-файл збережено з kolorami, фільтром i ramkami: {excel_path}")
    return excel_path

# Створення інтерфейсу користувача через Streamlit
st.title("Strona do przetwarzania plików HTML i generowania raportów Excel")
st.write("Załaduj plik HTML, aby utworzyć raport kolizji.")

# Завантаження файлу через Streamlit
uploaded_file = st.file_uploader("Wybierz plik HTML", type=["html"])

# Перевірка, чи файл завантажено
if uploaded_file is not None:
    # Додавання кнопки для перетворення файлу
    if st.button('Przekształć plik'):
        # Обробка HTML файлу
        excel_path = process_html_file(uploaded_file)
        
        # Показуємо кнопку для завантаження після обробки файлу
        st.success("Plik został przetworzony. Możesz teraz pobrać raport.")

        # Кнопка для завантаження Excel файлу
        st.download_button(
            label="Pobierz plik Excel",
            data=open(excel_path, "rb").read(),
            file_name="collisions_report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        # Видалення тимчасового файлу після завантаження
        os.remove(excel_path)