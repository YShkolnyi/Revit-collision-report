import pandas as pd
import streamlit as st
import os
from bs4 import BeautifulSoup
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.styles.borders import Border, Side
import io

# Функція для обробки файлу
def process_html_file(uploaded_file):

    # Завантаження HTML-файлу
    soup = BeautifulSoup(uploaded_file, "html.parser")

    # Знаходимо таблицю в HTML
    table = soup.find("table")

    # Отримуємо всі рядки таблиці
    rows = table.find_all("tr")

    data = []
    keys_0 = ["No."]
    keys_A = ["Workset A", "Category A", "Family A", "Type A", "ID A"]
    keys_B = ["Model B", "Category B", "Family B", "Type B", "ID B"]

    # Обробка рядків таблиці
    for row in rows[1:]:  # Пропускаємо заголовки
        raw_cols = row.find_all("td")
        cols = [raw_col.text.strip() for raw_col in raw_cols]

        # Перша комірка це лише номер
        separate_cols = {keys_0[0]: cols[0]}

        # Друга комірка це моя модель
        values_A = [item.strip() for item in cols[1].split(" : ")]
        dict_A = dict(zip(keys_A, values_A))
        separate_cols.update(dict_A)

        # Третя комірка це модель колізії
        values_B = [item.strip() for item in cols[2].split(" : ")]

        # Виправляємо невірно розбиті рядки
        if len(values_B) == 7:
            values_B[1] = " : ".join([values_B[1], values_B[2]])
            values_B[3] = " : ".join([values_B[3], values_B[4]])
            del values_B[2]
            del values_B[3]
            
        dict_B = dict(zip(keys_B, values_B))
        separate_cols.update(dict_B)

        # Додаємо стовпчик "Знак" і видаляємо лишні символи
        for key, value in separate_cols.copy().items():
            if "- Znak" in value:
                type_, sign = value.split("- Znak")
                if "A" in key:
                    separate_cols["Type A"] = type_.strip()
                    separate_cols["Sign A"] = sign.strip()
                else:
                    separate_cols["Type B"] = type_.strip()
                    separate_cols["Sign B"] = sign.strip()
            if "ID " in key:
                separate_cols[key] = value.replace('id ', '').strip()

        data.append(separate_cols)

    df = pd.DataFrame(data, columns=["No.", "||", "Workset A", "Category A", "Family A", "Type A", "Sign A", "ID A", "||",
                                     "Model B", "Category B", "Family B", "Type B", "Sign B", "ID B"])

    # Tworzenie obiektu BytesIO dla pliku Excel
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active

    # Zapisujemy dane DataFrame do arkusza
    for r_idx, row in enumerate(df.itertuples(index=False), start=1):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # Ustawienia wyglądu pliku Excel
    ws.freeze_panes = 'B2'  # Zamrożenie pierwszego wiersza
    ws.auto_filter.ref = ws.dimensions  # Filtry

    # Kolor nagłówków
    header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    header_font = Font(color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Zmiana kolorów wierszy
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for idx, cell in enumerate(row):
            if idx % 2 == 0:
                cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    # Dodanie obramowań
    border = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = border

    # Autodopasowanie szerokości kolumn
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2

    # Zapisujemy plik do obiektu BytesIO
    wb.save(output)
    output.seek(0)

    return output

# Strona Streamlit
st.title("Strona do przetwarzania plików HTML i generowania raportów Excel")
st.write("Załaduj plik HTML, aby utworzyć raport kolizji.")

uploaded_file = st.file_uploader("Wybierz plik HTML", type=["html"])

if uploaded_file is not None:
    if st.button('Przekształć plik'):
        excel_file = process_html_file(uploaded_file)

        st.success("Plik został przetworzony. Możesz teraz pobrać raport.")

        st.download_button(
            label="Pobierz plik Excel",
            data=excel_file,
            file_name="collisions_report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
